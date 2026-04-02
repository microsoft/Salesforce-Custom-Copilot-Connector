"""
Extract Owner/OwnerId from Salesforce API responses and ACL from
external item requests/responses in a deployment log file → Excel.

Usage:
    python extract_log_to_excel.py [log_file]

Defaults to the most recent deployment_*.log in the current directory.
"""

from __future__ import annotations

import glob
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


# ─── Helpers ──────────────────────────────────────────────────────────────────

LOG_LINE_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+ - \S+ - \w+ - (.*)$"
)

def strip_prefix(line: str) -> str:
    """Remove the log timestamp/level prefix and return just the message."""
    m = LOG_LINE_RE.match(line)
    return m.group(1) if m else line

def full_content(msg: str) -> str:
    """Strip log prefix from every line and return the joined content."""
    return "\n".join(strip_prefix(l) for l in msg.splitlines())

def read_log_messages(log_path: Path) -> list[str]:
    """
    Read the log file and collapse multi-line log entries into single strings.
    Each log entry starts with a timestamp.  Lines that don't start with a
    timestamp are continuation lines of the previous entry.
    """
    messages: list[str] = []
    buf: list[str] = []

    ts_re = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+")

    with log_path.open("r", encoding="utf-8", errors="replace") as fh:
        for raw in fh:
            line = raw.rstrip("\n")
            if ts_re.match(line):
                if buf:
                    messages.append("\n".join(buf))
                buf = [line]
            else:
                buf.append(line)
    if buf:
        messages.append("\n".join(buf))
    return messages


def extract_json_blob(messages: list[str], start_idx: int) -> tuple[dict | None, int]:
    """
    Starting from start_idx, collect consecutive lines that form a JSON object
    across multiple log entries and return the parsed dict + the index after it.
    """
    json_lines: list[str] = []
    i = start_idx
    brace_depth = 0
    in_json = False

    while i < len(messages):
        msg = messages[i]
        # strip log prefix from each line of a (possibly multi-line) entry
        content_lines = [strip_prefix(l) for l in msg.splitlines()]
        content = "\n".join(content_lines)

        for ch in content:
            if ch == "{":
                brace_depth += 1
                in_json = True
            elif ch == "}":
                brace_depth -= 1

        if in_json:
            json_lines.append(content)

        i += 1

        if in_json and brace_depth == 0:
            break

    if not json_lines:
        return None, start_idx + 1

    raw = "\n".join(json_lines)
    try:
        return json.loads(raw), i
    except json.JSONDecodeError:
        return None, i


# ─── Pass 1 – Salesforce raw records ──────────────────────────────────────────

SF_ITEM_RE   = re.compile(r"SALESFORCE ITEM \d+/\d+")
RAW_REC_RE   = re.compile(r"Raw Salesforce Record:")
OBJ_TYPE_RE  = re.compile(r"Object Type:\s*(.+)")
RECORD_ID_RE = re.compile(r"Record ID:\s*(.+)")

def parse_salesforce_records(messages: list[str]) -> dict[str, dict]:
    """
    Returns {record_id: {object_type, owner_name, owner_id, owner_role_id}}
    """
    records: dict[str, dict] = {}
    i = 0
    while i < len(messages):
        fc = full_content(messages[i])

        if SF_ITEM_RE.search(fc):
            obj_type = rec_id = ""
            # look ahead for Object Type and Record ID
            for j in range(i + 1, min(i + 8, len(messages))):
                c = full_content(messages[j])
                m = OBJ_TYPE_RE.search(c)
                if m:
                    obj_type = m.group(1).strip()
                m = RECORD_ID_RE.search(c)
                if m:
                    rec_id = m.group(1).strip()

            # find "Raw Salesforce Record:" then the JSON
            for j in range(i + 1, min(i + 12, len(messages))):
                c = full_content(messages[j])
                if RAW_REC_RE.search(c):
                    data, _ = extract_json_blob(messages, j + 1)
                    if data and rec_id:
                        owner_block = data.get("Owner") or {}
                        owner_name = ""
                        owner_role_id = ""
                        if isinstance(owner_block, dict):
                            owner_name = owner_block.get("Name", "")
                            user_role = owner_block.get("UserRole")
                            if isinstance(user_role, dict):
                                owner_role_id = user_role.get("Id", "")
                        records[rec_id] = {
                            "record_id": rec_id,
                            "object_type": obj_type,
                            "owner_name": owner_name,
                            "owner_id": data.get("OwnerId", ""),
                            "owner_role_id": owner_role_id,
                        }
                    break
        i += 1
    return records


# ─── Pass 2 – Item requests / responses (ACL) ─────────────────────────────────

ITEM_REQ_RE      = re.compile(r"(?<![A-Z ])ITEM REQUEST:\s*(\S+)")
SAMPLE_REQ_RE    = re.compile(r"SAMPLE ITEM REQUEST:\s*\S+\s+\(ID:\s*([^)]+)\)")
REQ_PAYLOAD_RE   = re.compile(r"Request Payload:")
RESPONSE_RE      = re.compile(r"^Response:$")

def _acl_list_to_str(acl: list) -> str:
    parts = []
    for entry in acl:
        if isinstance(entry, dict):
            parts.append(
                f"{entry.get('accessType','')}:{entry.get('type','')}:{entry.get('value','')}"
            )
    return " | ".join(parts)


def parse_item_requests(messages: list[str]) -> dict[str, dict]:
    """
    Returns {item_id: {request_acl_str, request_acl_raw,
                        response_acl_str, response_acl_raw, graph_status}}

    Two log formats handled:
      SAMPLE ITEM REQUEST  → has "Request Payload:" before JSON; has "Response:" block
      ITEM REQUEST         → JSON comes directly 2 messages after the header; no Response block
    """
    items: dict[str, dict] = {}
    i = 0
    while i < len(messages):
        fc = full_content(messages[i])

        item_id = None
        is_sample = False

        # ── Check SAMPLE first to avoid ITEM_REQ_RE capturing wrong text ──────
        m = SAMPLE_REQ_RE.search(fc)
        if m:
            item_id = m.group(1).strip()
            is_sample = True
        else:
            m = ITEM_REQ_RE.search(fc)
            if m:
                item_id = m.group(1).strip()

        if item_id:
            entry = items.setdefault(item_id, {
                "item_id": item_id,
                "request_acl_str": "",
                "request_acl_raw": [],
                "response_acl_str": "",
                "response_acl_raw": [],
                "graph_status": "sent",
            })

            if is_sample:
                # ── SAMPLE: "Request Payload:" appears before the JSON ──────────
                for j in range(i + 1, min(i + 8, len(messages))):
                    if REQ_PAYLOAD_RE.search(full_content(messages[j])):
                        payload, _ = extract_json_blob(messages, j + 1)
                        if payload and "acl" in payload:
                            entry["request_acl_raw"] = payload["acl"]
                            entry["request_acl_str"] = _acl_list_to_str(payload["acl"])
                        break

                # ── SAMPLE: "Response:" block is logged ──────────────────────────
                for j in range(i + 1, min(i + 35, len(messages))):
                    c = full_content(messages[j])
                    if RESPONSE_RE.search(c):
                        resp, _ = extract_json_blob(messages, j + 1)
                        if resp and "acl" in resp:
                            entry["response_acl_raw"] = resp["acl"]
                            entry["response_acl_str"] = _acl_list_to_str(resp["acl"])
                        entry["graph_status"] = "success"
                        break
                    if "Failed to load" in c and item_id in c:
                        entry["graph_status"] = "failed"
                        break

            else:
                # ── Regular ITEM REQUEST: JSON is the message right after "===" ──
                # Log sequence: [ITEM REQUEST] → [===] → [JSON payload] → [===\n]
                for j in range(i + 1, min(i + 5, len(messages))):
                    c = full_content(messages[j])
                    if c.lstrip().startswith("{"):
                        payload, _ = extract_json_blob(messages, j)
                        if payload and "acl" in payload:
                            entry["request_acl_raw"] = payload["acl"]
                            entry["request_acl_str"] = _acl_list_to_str(payload["acl"])
                        break

                # ── Regular ITEM REQUEST: no Response block; infer status ────────
                entry["graph_status"] = "success"
                for j in range(i + 1, min(i + 12, len(messages))):
                    c = full_content(messages[j])
                    if "Failed to load" in c and item_id in c:
                        entry["graph_status"] = "failed"
                        break

        i += 1
    return items


# ─── Build merged rows ────────────────────────────────────────────────────────

def build_rows(
    sf_records: dict[str, dict],
    item_requests: dict[str, dict],
) -> list[dict]:
    all_ids = sorted(set(sf_records) | set(item_requests))
    rows = []
    for rid in all_ids:
        sf  = sf_records.get(rid, {})
        req = item_requests.get(rid, {})
        rows.append({
            "Record ID":         rid,
            "Object Type":       sf.get("object_type", ""),
            "Owner Name":        sf.get("owner_name", ""),
            "Owner ID":          sf.get("owner_id", ""),
            "Owner Role ID":     sf.get("owner_role_id", ""),
            "Request ACL":       req.get("request_acl_str", ""),
            "Response ACL":      req.get("response_acl_str", ""),
            "Graph Status":      req.get("graph_status", ""),
        })
    return rows


# ─── Excel writer ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
FAILED_FILL  = PatternFill("solid", fgColor="FFCCCC")
SUCCESS_FILL = PatternFill("solid", fgColor="CCFFCC")

OBJ_COLORS = {
    "Account":     "FFF2CC",
    "Contact":     "D9EAD3",
    "Lead":        "FCE5CD",
    "Opportunity": "CFE2F3",
    "Case":        "EAD1DC",
}

def write_excel(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Records ──────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Records"

    headers = list(rows[0].keys()) if rows else []
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        obj_fill_hex = OBJ_COLORS.get(row.get("Object Type", ""), None)
        if obj_fill_hex:
            obj_fill = PatternFill("solid", fgColor=obj_fill_hex)
        else:
            obj_fill = None

        for col, key in enumerate(headers, 1):
            val = row[key]
            cell = ws_all.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if key == "Object Type" and obj_fill:
                cell.fill = obj_fill
            elif key == "Graph Status":
                if val == "success":
                    cell.fill = SUCCESS_FILL
                elif val == "failed":
                    cell.fill = FAILED_FILL
                elif fill:
                    cell.fill = fill
            elif fill:
                cell.fill = fill

    # auto-width
    col_widths = {"Record ID": 22, "Object Type": 14, "Owner Name": 22,
                  "Owner ID": 22, "Owner Role ID": 22,
                  "Request ACL": 60, "Response ACL": 60, "Graph Status": 13}
    for col, h in enumerate(headers, 1):
        ws_all.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 18)

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions

    # ── Sheet 2: Per-Object-Type summary sheets ───────────────────────────────
    by_type: dict[str, list[dict]] = {}
    for row in rows:
        by_type.setdefault(row["Object Type"] or "Unknown", []).append(row)

    for obj_type, obj_rows in sorted(by_type.items()):
        ws = wb.create_sheet(title=obj_type[:31])
        hex_color = OBJ_COLORS.get(obj_type, "E8E8E8")
        hdr_fill = PatternFill("solid", fgColor=hex_color)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font  = Font(bold=True, size=10)
            cell.fill  = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, row in enumerate(obj_rows, 2):
            for col, key in enumerate(headers, 1):
                val = row[key]
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if key == "Graph Status":
                    if val == "success":
                        cell.fill = SUCCESS_FILL
                    elif val == "failed":
                        cell.fill = FAILED_FILL

        for col, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 18)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 3: ACL Summary (unique ACL patterns) ────────────────────────────
    ws_acl = wb.create_sheet(title="ACL Summary")
    acl_headers = ["Object Type", "ACL Type", "Access Type", "Value", "Count"]
    for col, h in enumerate(acl_headers, 1):
        cell = ws_acl.cell(row=1, column=col, value=h)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    from collections import Counter
    acl_counter: Counter = Counter()
    for row in rows:
        for entry_str in row["Response ACL"].split(" | "):
            if not entry_str:
                continue
            parts = entry_str.split(":")
            if len(parts) == 3:
                access_type, acl_type, value = parts
                acl_counter[(row["Object Type"], acl_type.strip(), access_type.strip(), value.strip())] += 1

    for row_idx, ((obj_type, acl_type, access_type, value), count) in enumerate(
        sorted(acl_counter.items()), 2
    ):
        ws_acl.cell(row=row_idx, column=1, value=obj_type)
        ws_acl.cell(row=row_idx, column=2, value=acl_type)
        ws_acl.cell(row=row_idx, column=3, value=access_type)
        ws_acl.cell(row=row_idx, column=4, value=value)
        ws_acl.cell(row=row_idx, column=5, value=count)

    for col, w in zip(range(1, 6), [16, 16, 14, 42, 8]):
        ws_acl.column_dimensions[get_column_letter(col)].width = w

    wb.save(out_path)
    print(f"✅ Excel saved to: {out_path}")
    print(f"   Sheets: All Records ({len(rows)} rows) + per-object sheets + ACL Summary")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    here = Path(__file__).parent

    if len(sys.argv) > 1:
        log_path = Path(sys.argv[1])
    else:
        candidates = sorted(here.glob("deployment_*.log"), reverse=True)
        if not candidates:
            print("No deployment_*.log found in current directory.")
            sys.exit(1)
        log_path = candidates[0]

    print(f"📂 Parsing: {log_path.name}  ({log_path.stat().st_size / 1_048_576:.1f} MB)")

    print("   Reading log messages …")
    messages = read_log_messages(log_path)
    print(f"   {len(messages):,} log entries loaded")

    print("   Extracting Salesforce records …")
    sf_records = parse_salesforce_records(messages)
    print(f"   {len(sf_records):,} Salesforce records found")

    print("   Extracting item requests / ACLs …")
    item_requests = parse_item_requests(messages)
    print(f"   {len(item_requests):,} external item requests found")

    rows = build_rows(sf_records, item_requests)
    print(f"   {len(rows):,} merged rows")

    out_path = here / (log_path.stem + "_extract_v2.xlsx")
    write_excel(rows, out_path)


if __name__ == "__main__":
    main()
